from openpyxl import Workbook

wb = Workbook()
ws = wb.active

columns = ["Show", "Duplicate Record"]
ws.append(columns)


def write_fox_shows(shows, file='show_dupes.xlsx'):
    for row in zip(shows):
        ws.append(row)
    wb.save(file)


def write_collection_if_dupe(fox_last_four_shows, collection_shows, collection, file='show_dupes.xlsx'):
    dupes = list(set(fox_last_four_shows).intersection(collection_shows))
    if dupes:
        for fox_show in fox_last_four_shows:
            if fox_show in collection_shows:
                _write_dupe_collection_cells(collection=collection, index=fox_last_four_shows.index(fox_show), file=file)


def _write_dupe_collection_cells(collection, index, file):
    # Lower it below Column name
    cell_index = str(index + 2)
    old_collection = ws['B%s' % cell_index].value
    if old_collection is None:
        ws['B%s' % cell_index] = collection
    # Multiple collections in same cell
    else:
        ws['B%s' % cell_index] = old_collection + ', ' + str(collection)
    wb.save(file)